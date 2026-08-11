"""
Общая логика конвертера спецификаций — используется и CLI (convert.py),
и веб-бэкендом (backend/main.py).

Содержит:
  * чтение текста из PDF (из файла или из потока байтов);
  * разбор текста через Gemini в структурированный JSON (с ретраями);
  * сборку и форматирование Excel-книги из уже разобранных данных.
"""

import os
import re
import json
import time
import io

import pypdf
from pypdf import PdfReader
from dotenv import load_dotenv
from google import genai
from google.genai import types
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Игнорируем мелкие ошибки разметки внутри PDF из AutoCAD
pypdf.PdfReader.strict = False

load_dotenv()

# Модель Gemini для парсинга (можно переопределить через переменную окружения)
GEMINI_MODEL = os.getenv("GEMINI_MODEL", "gemini-3.6-flash")

# Ограничения ширины колонок Excel
MAX_COLUMN_WIDTH = 60
MIN_COLUMN_WIDTH = 10

# Порядок и подписи колонок итоговой таблицы: (заголовок в Excel, ключ в JSON от ИИ)
ITEM_COLUMNS = [
    ("Раздел", "section"),
    ("Позиция в спецификации", "pos"),
    ("Наименование и техническая характеристика", "name"),
    ("Тип, марка, обозначение", "type_code"),
    ("Код продукции", "product_code"),
    ("Поставщик", "supplier"),
    ("Ед. изм.", "unit"),
    ("Кол-во", "quantity"),
    ("Примечание", "note"),
]

# Колонки уровня документа (идут перед позиционными)
DOC_COLUMNS = ["Объект", "Шифр документа", "Система"]


def all_headers():
    """Полный упорядоченный список заголовков колонок Excel."""
    return DOC_COLUMNS + [header for header, _ in ITEM_COLUMNS]

# Схема JSON-вывода Gemini (Structured Output)
SPEC_SCHEMA = {
    "type": "OBJECT",
    "properties": {
        "doc_number": {"type": "STRING", "description": "Шифр/номер документа из штампа"},
        "object_name": {"type": "STRING", "description": "Наименование объекта/сооружения из штампа (например 'Склад крупнодробленой руды'), без названия системы и без стадии"},
        "system_name": {"type": "STRING", "description": "Наименование системы/раздела"},
        "items": {
            "type": "ARRAY",
            "items": {
                "type": "OBJECT",
                "properties": {
                    "section": {"type": "STRING", "description": "Раздел спецификации, под которым идёт позиция: например 'Оборудование', 'Кабели и провода', 'Изделия и материалы'. Если раздел не указан — пустая строка."},
                    "pos": {"type": "STRING", "description": "Позиция/Марка"},
                    "name": {"type": "STRING", "description": "Наименование и техническая характеристика"},
                    "type_code": {"type": "STRING", "description": "Тип, марка, обозначение документа, опросного листа"},
                    "product_code": {"type": "STRING", "description": "Код продукции"},
                    "supplier": {"type": "STRING", "description": "Поставщик/Изготовитель"},
                    "unit": {"type": "STRING", "description": "Единица измерения"},
                    "quantity": {"type": "NUMBER", "description": "Количество"},
                    "weight_kg": {"type": "NUMBER", "description": "Масса 1 ед, кг"},
                    "note": {"type": "STRING", "description": "Примечание"}
                },
                "required": ["pos", "name", "quantity", "unit"]
            }
        }
    },
    "required": ["doc_number", "items"]
}

SYSTEM_PROMPT = """
Ты — инженерный ассистент. Твоя задача — извлечь данные спецификации оборудования,
изделий и материалов из текста чертежа ГОСТ и сгруппировать их согласно структуре JSON.

Правила:
- Пропускай заголовки таблиц и служебные надписи (например «Внимание! Перед заказом...»).
- Не придумывай данные, которых нет в тексте.
- Спецификация обычно разбита на разделы («Оборудование», «Кабели и провода»,
  «Изделия и материалы» и т.п.), и нумерация позиций в каждом разделе начинается заново.
  Для КАЖДОЙ позиции указывай её раздел в поле "section". Если позиция явно не под разделом —
  оставь "section" пустым.
- Если внутри одной ячейки текст перенесён на несколько строк, объединяй его через пробел,
  не склеивай слова вместе (например «Пожарной Автоматики», а не «ПожарнойАвтоматики»).
- Поле "weight_kg" (масса) заполняй только если масса реально указана в тексте; если её нет —
  не ставь 0, оставь поле пустым (не указывай).
- В поле "object_name" укажи наименование объекта/сооружения из штампа (то, ГДЕ монтируется
  система), например «Склад крупнодробленой руды» — без названия системы и без стадии.
- ЧИСЛА (количество и масса) переписывай ПРЕДЕЛЬНО ТОЧНО — все цифры до единой, ровно как
  в тексте. Не пропускай цифры, не округляй, не отбрасывай разряды. Например, если в тексте
  «1053» — это 1053, а не 105. Количество бери из колонки «Количество», а не из описания
  (в описании может быть длина одной бухты, напр. «50 м» — это НЕ количество).
"""


def get_api_key():
    """Возвращает ключ Gemini из окружения или бросает понятную ошибку."""
    key = os.getenv("GEMINI_API_KEY", "").strip()
    if not key:
        raise RuntimeError(
            "Не найден GEMINI_API_KEY. Создайте .env на основе .env.example "
            "и вставьте ключ (https://aistudio.google.com/apikey)."
        )
    return key


def get_client():
    """Создаёт клиент Gemini."""
    return genai.Client(api_key=get_api_key())


def extract_text_from_pdf(source):
    """Извлекает цифровой текст из векторного PDF.

    source — путь к файлу (str) ИЛИ файловый объект / поток байтов (BytesIO).
    """
    reader = PdfReader(source)
    full_text = []
    for idx, page in enumerate(reader.pages):
        text = page.extract_text()
        if text:
            full_text.append(f"--- СТРАНИЦА {idx + 1} ---\n{text}")
    return "\n\n".join(full_text)


def parse_specification_with_ai(text_content, client=None, max_retries=5):
    """Отправляет текст в Gemini и получает структурированный JSON.

    При временных ошибках сервера (503 — модель перегружена, 429 — лимит)
    повторяет запрос с экспоненциальной задержкой.
    """
    if client is None:
        client = get_client()

    for attempt in range(1, max_retries + 1):
        try:
            response = client.models.generate_content(
                model=GEMINI_MODEL,
                contents=[
                    types.Part.from_text(text=f"Текст спецификации:\n\n{text_content}"),
                ],
                config=types.GenerateContentConfig(
                    system_instruction=SYSTEM_PROMPT,
                    response_mime_type="application/json",
                    response_schema=SPEC_SCHEMA,
                    temperature=0  # максимально детерминированно — меньше случайных промахов
                )
            )
            return json.loads(response.text)
        except Exception as e:
            code = getattr(e, "code", None)
            is_transient = code in (429, 503) or "503" in str(e) or "429" in str(e)
            if not is_transient or attempt == max_retries:
                raise
            time.sleep(2 ** attempt)  # 2, 4, 8, 16 секунд


def full_shifr(filename):
    """Полный шифр документа = имя файла без расширения (в нём есть и суффикс ревизии, напр. 'C2')."""
    return os.path.splitext(filename or "")[0].strip()


def system_code(shifr):
    """Аббревиатура системы из шифра: ПОСЛЕДНИЙ сегмент из 2-5 заглавных букв перед числом.
    Напр. 'MOF3-UN-300000-INS-AFS-0021' → 'AFS' (а не 'UN')."""
    codes = re.findall(r'([A-Z]{2,5})-\d', shifr or "")
    return codes[-1] if codes else ""


def items_to_rows(res):
    """Превращает результат разбора одного файла в плоские строки.

    res — словарь: {"filename", "doc_number", "system_name", "object_name", "items"}
    Колонки уровня документа:
      Объект        — наименование объекта (object_name, иначе system_name)
      Шифр документа — полный шифр из имени файла (с суффиксом ревизии)
      Система        — аббревиатура (AFS и т.п.), извлечённая из шифра
    """
    filename = res.get("filename", "")
    shifr = full_shifr(filename)
    obj = (res.get("object_name") or res.get("system_name") or "").strip()
    sys_code = system_code(shifr) or (res.get("system_name") or "")

    rows = []
    for item in res.get("items", []):
        row = {
            "Объект": obj,
            "Шифр документа": shifr,
            "Система": sys_code,
        }
        for header, key in ITEM_COLUMNS:
            value = item.get(key, "")
            if value is None:
                value = ""
            row[header] = value
        rows.append(row)
    return rows


def safe_sheet_name(filename, used_names):
    """Валидное и уникальное имя листа Excel (макс. 31 символ)."""
    name = os.path.splitext(filename)[0]
    name = re.sub(r'[\[\]:*?/\\]', '_', name)
    name = name[:31]

    original = name
    counter = 1
    while name in used_names:
        suffix = f"_{counter}"
        name = original[:31 - len(suffix)] + suffix
        counter += 1

    used_names.add(name)
    return name


def _write_sheet(worksheet, headers, rows):
    """Пишет заголовки и строки на лист и применяет форматирование (без pandas)."""
    # Шапка
    worksheet.append(headers)

    # Данные — в порядке headers
    for row in rows:
        worksheet.append([row.get(h, "") for h in headers])

    # --- Стилизация шапки ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # --- Ширина колонок по самому длинному значению ---
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in rows:
            value = row.get(header, "")
            value_len = len(str(value)) if value is not None else 0
            if value_len > max_len:
                max_len = value_len
        width = min(max(max_len + 2, MIN_COLUMN_WIDTH), MAX_COLUMN_WIDTH)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Перенос текста в данных ---
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


# Колонки итоговой сводки по наименованиям
SUMMARY_COLUMNS = [
    "Наименование и техническая характеристика",
    "Тип, марка, обозначение",
    "Код продукции",
    "Ед. изм.",
    "Итого количество",
    "Кол-во объектов",
]


def _norm(s):
    """Нормализует строку: нижний регистр, схлопнутые пробелы."""
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def summarize_by_name(all_rows):
    """Сводит позиции и суммирует количество по всем спецификациям.

    Один и тот же товар в разных спецификациях часто назван по-разному, но код у него
    один. Поэтому группируем по КОДУ (Код продукции, иначе Тип/марка), а если кода нет —
    по нормализованному наименованию. Единицу измерения держим в ключе (не мешаем шт. и м).
    """
    groups = {}
    for r in all_rows:
        name = (r.get("Наименование и техническая характеристика") or "").strip()
        if not name:
            continue
        unit = (r.get("Ед. изм.") or "").strip()
        code = (r.get("Код продукции") or r.get("Тип, марка, обозначение") or "").strip()
        # ключ: по коду, если он есть; иначе по имени
        key = (("c:" + _norm(code)) if code else ("n:" + _norm(name)), _norm(unit))

        g = groups.get(key)
        if g is None:
            g = {"name": name, "unit": unit, "qty": 0,
                 "type_code": r.get("Тип, марка, обозначение") or "",
                 "product_code": r.get("Код продукции") or "",
                 "objects": set()}
            groups[key] = g
        qty = r.get("Кол-во")
        if isinstance(qty, (int, float)):
            g["qty"] += qty
        # показываем самое длинное (обычно самое подробное) наименование
        if len(name) > len(g["name"]):
            g["name"] = name
        if not g["type_code"]:
            g["type_code"] = r.get("Тип, марка, обозначение") or ""
        if not g["product_code"]:
            g["product_code"] = r.get("Код продукции") or ""
        g["objects"].add(r.get("Шифр документа") or "")

    rows = []
    for g in groups.values():
        qty = g["qty"]
        rows.append({
            "Наименование и техническая характеристика": g["name"],
            "Тип, марка, обозначение": g["type_code"],
            "Код продукции": g["product_code"],
            "Ед. изм.": g["unit"],
            "Итого количество": int(qty) if float(qty).is_integer() else qty,
            "Кол-во объектов": len(g["objects"]),
        })
    rows.sort(key=lambda x: x["Наименование и техническая характеристика"].lower())
    return rows


def build_workbook(file_results, output=None):
    """Собирает Excel-книгу из результатов разбора (на openpyxl, без pandas).

    file_results — список словарей:
        {"filename", "doc_number", "system_name", "object_name", "items": [...]}

    output — путь к файлу (str) или None. Если None — вернёт BytesIO с готовой книгой.
    """
    headers = all_headers()
    all_rows = []
    per_file = []  # (filename, rows)

    for res in file_results:
        rows = items_to_rows(res)
        if rows:
            all_rows.extend(rows)
            per_file.append((res.get("filename", "лист"), rows))

    if not all_rows:
        raise ValueError("Нет данных для записи в Excel.")

    wb = Workbook()

    # 1) Лист «Итого по наименованиям» — сумма количеств по всем спецификациям
    summary_by_name_name = "Итого по наименованиям"
    ws_total = wb.active
    ws_total.title = summary_by_name_name
    _write_sheet(ws_total, SUMMARY_COLUMNS, summarize_by_name(all_rows))

    # 2) Сводный лист со всеми позициями
    summary_sheet_name = "Сводная спецификация"
    ws_summary = wb.create_sheet(title=summary_sheet_name)
    _write_sheet(ws_summary, headers, all_rows)

    # 3) Лист на каждый файл
    used_names = {summary_by_name_name, summary_sheet_name}
    for filename, rows in per_file:
        sheet_name = safe_sheet_name(filename, used_names)
        ws = wb.create_sheet(title=sheet_name)
        _write_sheet(ws, headers, rows)

    target = output if output is not None else io.BytesIO()
    wb.save(target)

    if output is None:
        target.seek(0)
    return target
